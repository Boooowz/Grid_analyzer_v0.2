# -*- coding: utf-8 -*-
"""
Εργαλείο σύνοψης ωρών εργασίας από Excel ΕΡΓΑΝΗ/προγράμματος απασχόλησης.

Διαβάζει αρχείο .xlsx με στήλες:
ΑΦΜ, Όνομα, Επώνυμο, Ημ/νια, Διάλειμμα, Απασχόληση
και εξάγει συνοπτικά ανά εργαζόμενο:
- Ημέρες εργασίας
- Ώρες εργασίας όπως προκύπτουν από το ωράριο
- Νυχτερινές ώρες (22:00-06:00)
- Ημέρες εργασίας σε Κυριακή
- Ώρες εργασίας Κυριακής
- Νυχτερινές ώρες Κυριακής

Κρίσιμοι κανόνες:
1. Γραμμές που περιέχουν ένδειξη άδειας, π.χ. «Κανονική άδεια», εξαιρούνται πλήρως
   από ημέρες και ώρες, ακόμη κι αν στη γραμμή υπάρχει τυπικά ωράριο.
2. Οι γραμμές «ΜΗ ΕΡΓΑΣΙΑ» και «ΑΝΑΠΑΥΣΗ/ΡΕΠΟ» δεν υπολογίζονται ως εργασία.
3. Το διάλειμμα τύπου «Εντός 10», «Εντός 15», «Εντός 20» δεν αφαιρείται από τις ώρες,
   επειδή θεωρείται ότι περιλαμβάνεται μέσα στο δηλωμένο ωράριο εργασίας.
   Καταγράφεται μόνο ενημερωτικά στο αποτέλεσμα.

Δεν υπολογίζει αποδοχές ή προσαυξήσεις. Υπολογίζει μόνο ημέρες/ώρες.
"""

from __future__ import annotations

import csv
import re
import sys
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, date, time, timedelta
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Set, Tuple

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.datetime import from_excel
except ImportError as exc:
    print("Λείπει η βιβλιοθήκη openpyxl. Τρέξτε πρώτα το run_app.bat ή εγκαταστήστε: pip install openpyxl")
    raise


NIGHT_START = time(22, 0)
NIGHT_END = time(6, 0)
TIME_INTERVAL_RE = re.compile(r"(\d{1,2})\s*[:.]\s*(\d{2})\s*[-–—]\s*(\d{1,2})\s*[:.]\s*(\d{2})")


@dataclass
class DetailRow:
    source_row: int
    afm: str
    full_name: str
    work_date: date
    break_minutes: float
    apasxolisi: str
    start_dt: datetime
    end_dt: datetime
    gross_hours: float
    net_hours: float
    gross_night_hours: float
    net_night_hours: float
    gross_sunday_hours: float
    net_sunday_hours: float
    gross_sunday_night_hours: float
    net_sunday_night_hours: float
    sunday_dates: List[date] = field(default_factory=list)
    holiday_dates: List[date] = field(default_factory=list)
    note: str = ""


@dataclass
class ExcludedRow:
    source_row: int
    afm: str
    full_name: str
    work_date: Optional[date]
    apasxolisi: str
    reason: str


@dataclass
class EmployeeSummary:
    afm: str
    first_name: str
    last_name: str
    full_name: str
    work_dates: set = field(default_factory=set)
    sunday_dates: set = field(default_factory=set)
    holiday_dates: set = field(default_factory=set)
    excluded_leave_dates: set = field(default_factory=set)
    excluded_leave_rows: int = 0
    excluded_non_work_rows: int = 0
    gross_hours: float = 0.0
    break_hours: float = 0.0
    total_hours: float = 0.0
    night_hours: float = 0.0
    sunday_hours: float = 0.0
    sunday_night_hours: float = 0.0
    work_rows: int = 0
    warnings: List[str] = field(default_factory=list)


def norm_text(value) -> str:
    """Normalize Greek/Latin header text for safer matching."""
    if value is None:
        return ""
    txt = str(value).strip().lower()
    txt = unicodedata.normalize("NFD", txt)
    txt = "".join(ch for ch in txt if unicodedata.category(ch) != "Mn")
    txt = re.sub(r"\s+", "", txt)
    txt = txt.replace("/", "").replace("-", "").replace("_", "")
    return txt


def is_leave_row(apasholisi_text: str) -> bool:
    """Return True for rows that represent leave and must not be counted as work time."""
    normalized = norm_text(apasholisi_text)
    leave_tokens = (
        "αδεια",       # άδεια / αδεια, κανονική άδεια κ.λπ.
        "adeia",       # πιθανή λατινική καταχώρηση
        "leave",       # πιθανή αγγλική καταχώρηση
    )
    return any(token in normalized for token in leave_tokens)


def is_non_work_row(apasholisi_text: str) -> bool:
    """Rows that are explicitly not work."""
    normalized = norm_text(apasholisi_text)
    non_work_tokens = (
        "μηεργασια",
        "αναπαυσηρεπο",
        "αναπαυση",
        "ρεπο",
        "repo",
        "rest",
    )
    return any(token in normalized for token in non_work_tokens)


def find_header_map(headers: List[str]) -> Dict[str, Optional[int]]:
    aliases = {
        "afm": {"αφμ", "afm"},
        "first_name": {"ονομα", "name", "firstname", "first"},
        "last_name": {"επωνυμο", "surname", "lastname", "last"},
        "date": {"ημνια", "ημερομηνια", "date", "hmerominia"},
        "employment": {"απασχοληση", "εργασια", "employment", "work"},
    }
    optional_aliases = {
        "break": {"διαλειμμα", "break", "breaktime"},
    }
    normalized = [norm_text(h) for h in headers]
    result: Dict[str, Optional[int]] = {}
    for key, candidates in aliases.items():
        for idx, h in enumerate(normalized):
            if h in candidates:
                result[key] = idx
                break
    missing = [key for key in aliases if key not in result]
    if missing:
        pretty = {
            "afm": "ΑΦΜ",
            "first_name": "Όνομα",
            "last_name": "Επώνυμο",
            "date": "Ημ/νια",
            "employment": "Απασχόληση",
        }
        raise ValueError("Δεν βρέθηκαν οι απαραίτητες στήλες: " + ", ".join(pretty[m] for m in missing))
    for key, candidates in optional_aliases.items():
        result[key] = None
        for idx, h in enumerate(normalized):
            if h in candidates:
                result[key] = idx
                break
    return result


def parse_date(value) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date()
        except Exception:
            return None
    text = str(value).strip()
    text = text.replace("\\", "/").replace(".", "/").replace("-", "/")
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def parse_time(h: str, m: str) -> time:
    hour = int(h)
    minute = int(m)
    if hour == 24 and minute == 0:
        hour = 0
    if not (0 <= hour <= 23 and 0 <= minute <= 59):
        raise ValueError(f"Μη έγκυρη ώρα: {h}:{m}")
    return time(hour, minute)


def parse_break_minutes(value) -> float:
    """Parse break values such as 'Εντός 10', 'Εντός 20', '00:15', Excel time, or numeric minutes."""
    if value is None or str(value).strip() == "":
        return 0.0
    if isinstance(value, datetime):
        return value.hour * 60 + value.minute + value.second / 60
    if isinstance(value, time):
        return value.hour * 60 + value.minute + value.second / 60
    if isinstance(value, (int, float)):
        # Excel time is usually fraction of a day; plain values >= 1 are treated as minutes.
        if 0 < float(value) < 1:
            return float(value) * 24 * 60
        return float(value)
    text = str(value).strip()
    normalized = norm_text(text)
    if normalized in {"οχι", "none", "χωρις", "χωριςδιαλειμμα", "nodata", "nan"}:
        return 0.0
    hm = re.search(r"(\d{1,2})\s*[:.]\s*(\d{2})", text)
    if hm:
        return int(hm.group(1)) * 60 + int(hm.group(2))
    nums = re.findall(r"\d+(?:[,.]\d+)?", text)
    if nums:
        return float(nums[0].replace(",", "."))
    return 0.0


def extract_intervals(apasholisi_text: str, base_date: date) -> List[Tuple[datetime, datetime]]:
    intervals: List[Tuple[datetime, datetime]] = []
    if not apasholisi_text:
        return intervals
    for match in TIME_INTERVAL_RE.finditer(str(apasholisi_text)):
        sh, sm, eh, em = match.groups()
        st = parse_time(sh, sm)
        et = parse_time(eh, em)
        start_dt = datetime.combine(base_date, st)
        end_dt = datetime.combine(base_date, et)
        # Βάρδια που περνάει τα μεσάνυχτα, π.χ. 21:00-01:00 ή 18:00-00:00.
        if end_dt <= start_dt:
            end_dt += timedelta(days=1)
        intervals.append((start_dt, end_dt))
    return intervals


def overlap_hours(start: datetime, end: datetime, win_start: datetime, win_end: datetime) -> float:
    latest = max(start, win_start)
    earliest = min(end, win_end)
    seconds = (earliest - latest).total_seconds()
    if seconds <= 0:
        return 0.0
    return seconds / 3600.0


def iter_dates(start: date, end: date) -> Iterable[date]:
    current = start
    while current <= end:
        yield current
        current += timedelta(days=1)


def night_hours(start: datetime, end: datetime) -> float:
    total = 0.0
    # Ψάχνουμε γύρω από το διάστημα ώστε να πιάνονται και παράθυρα που ξεκινούν την προηγούμενη μέρα.
    for d in iter_dates(start.date() - timedelta(days=1), end.date()):
        ns = datetime.combine(d, NIGHT_START)
        ne = datetime.combine(d + timedelta(days=1), NIGHT_END)
        total += overlap_hours(start, end, ns, ne)
    return total


def sunday_overlap(start: datetime, end: datetime,
                   holiday_dates: Optional[Set[date]] = None) -> Tuple[float, List[date]]:
    hd = holiday_dates or set()
    hours = 0.0
    dates: List[date] = []
    for d in iter_dates(start.date(), end.date()):
        # Python: Monday=0 ... Sunday=6. Οι αργίες μετριούνται όπως η Κυριακή.
        if d.weekday() == 6 or d in hd:
            ss = datetime.combine(d, time(0, 0))
            se = datetime.combine(d + timedelta(days=1), time(0, 0))
            h = overlap_hours(start, end, ss, se)
            if h > 0:
                hours += h
                dates.append(d)
    return hours, dates


def sunday_night_hours(start: datetime, end: datetime,
                       holiday_dates: Optional[Set[date]] = None) -> float:
    hd = holiday_dates or set()
    total = 0.0
    for d in iter_dates(start.date(), end.date()):
        if d.weekday() == 6 or d in hd:
            # Νυχτερινές ώρες μέσα στην Κυριακή/αργία: 00:00-06:00 και 22:00-24:00.
            total += overlap_hours(start, end, datetime.combine(d, time(0, 0)), datetime.combine(d, NIGHT_END))
            total += overlap_hours(start, end, datetime.combine(d, NIGHT_START), datetime.combine(d + timedelta(days=1), time(0, 0)))
    return total


def _orthodox_easter(year: int) -> date:
    # Meeus Julian algorithm + 13 days shift για το Γρηγοριανό (ισχύει 1900-2099).
    a = year % 4
    b = year % 7
    c = year % 19
    d_ = (19 * c + 15) % 30
    e = (2 * a + 4 * b - d_ + 34) % 7
    month = (d_ + e + 114) // 31
    day = ((d_ + e + 114) % 31) + 1
    julian = date(year, month, day)
    return julian + timedelta(days=13)


def greek_public_holidays(year: int) -> Dict[date, str]:
    fixed: Dict[date, str] = {
        date(year, 1, 1): "Πρωτοχρονιά",
        date(year, 1, 6): "Θεοφάνεια",
        date(year, 3, 25): "25η Μαρτίου",
        date(year, 5, 1): "Πρωτομαγιά",
        date(year, 8, 15): "Δεκαπενταύγουστος",
        date(year, 10, 28): "28η Οκτωβρίου",
        date(year, 12, 25): "Χριστούγεννα",
        date(year, 12, 26): "Σύναξη Θεοτόκου",
    }
    easter = _orthodox_easter(year)
    movable: Dict[date, str] = {
        easter - timedelta(days=48): "Καθαρά Δευτέρα",
        easter - timedelta(days=2): "Μ. Παρασκευή",
        easter + timedelta(days=1): "Δευτέρα Πάσχα",
        easter + timedelta(days=50): "Αγίου Πνεύματος",
    }
    return {**fixed, **movable}


def round_hours(value: float) -> float:
    return round(value + 1e-9, 2)


def leave_note(count: int) -> str:
    if count <= 0:
        return ""
    if count == 1:
        return "Εξαιρέθηκε 1 γραμμή άδειας. "
    return f"Εξαιρέθηκαν {count} γραμμές άδειας. "


def holiday_note(holiday_dates_set) -> str:
    if not holiday_dates_set:
        return ""
    parts = ", ".join(d.strftime("%d/%m") for d in sorted(holiday_dates_set))
    return f"Αργίες ως Κυριακή: {parts}. "


def analyze_excel(input_path: str | Path,
                  holiday_dates: Optional[Set[date]] = None) -> Tuple[List[EmployeeSummary], List[DetailRow], List[ExcludedRow], List[str]]:
    input_path = Path(input_path)
    hd = holiday_dates or set()
    wb = load_workbook(input_path, data_only=True)
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    header_map = find_header_map(headers)

    summaries: Dict[Tuple[str, str, str], EmployeeSummary] = {}
    details: List[DetailRow] = []
    excluded_rows: List[ExcludedRow] = []
    general_warnings: List[str] = []

    for r in range(2, ws.max_row + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None or str(v).strip() == "" for v in values):
            continue

        afm = str(values[header_map["afm"]] or "").strip()
        first_name = str(values[header_map["first_name"]] or "").strip()
        last_name = str(values[header_map["last_name"]] or "").strip()
        full_name = (first_name + " " + last_name).strip()
        dt = parse_date(values[header_map["date"]])
        apasholisi = str(values[header_map["employment"]] or "").strip()
        break_value = values[header_map["break"]] if header_map.get("break") is not None else None
        break_minutes = parse_break_minutes(break_value)

        if not afm and not full_name:
            continue
        key = (afm, first_name, last_name)
        if key not in summaries:
            summaries[key] = EmployeeSummary(afm=afm, first_name=first_name, last_name=last_name, full_name=full_name)
        summary = summaries[key]

        if dt is None:
            if apasholisi:
                msg = f"Γραμμή {r}: δεν αναγνωρίστηκε ημερομηνία ({values[header_map['date']]})."
                summary.warnings.append(msg)
                general_warnings.append(msg)
            continue

        # Η άδεια δεν είναι πραγματική εργασία.
        # Αγνοείται πλήρως ακόμη κι αν το πεδίο «Απασχόληση» περιέχει και ωράριο.
        if is_leave_row(apasholisi):
            summary.excluded_leave_dates.add(dt)
            summary.excluded_leave_rows += 1
            excluded_rows.append(ExcludedRow(r, afm, full_name, dt, apasholisi, "Άδεια - εξαιρέθηκε πλήρως"))
            continue

        # Ρητές γραμμές μη εργασίας/ρεπό δεν υπολογίζονται και δεν δημιουργούν προειδοποίηση.
        if is_non_work_row(apasholisi):
            summary.excluded_non_work_rows += 1
            excluded_rows.append(ExcludedRow(r, afm, full_name, dt, apasholisi, "Μη εργασία/ρεπό"))
            continue

        try:
            intervals = extract_intervals(apasholisi, dt)
        except Exception as exc:
            msg = f"Γραμμή {r}: σφάλμα ανάγνωσης ωραρίου '{apasholisi}': {exc}"
            summary.warnings.append(msg)
            general_warnings.append(msg)
            continue

        if not intervals and "ΕΡΓΑΣΙΑ" in apasholisi.upper():
            msg = f"Γραμμή {r}: βρέθηκε ένδειξη εργασίας χωρίς αναγνωρίσιμο ωράριο: '{apasholisi}'."
            summary.warnings.append(msg)
            general_warnings.append(msg)
            continue

        total_gross_row_hours = sum((end_dt - start_dt).total_seconds() / 3600.0 for start_dt, end_dt in intervals)
        if total_gross_row_hours <= 0:
            continue
        # Το διάλειμμα ΔΕΝ αφαιρείται από τις ώρες εργασίας.
        # Θεωρείται ότι περιλαμβάνεται στο δηλωμένο ωράριο και κρατείται μόνο ενημερωτικά.
        break_hours_row = min(break_minutes / 60.0, total_gross_row_hours)

        for start_dt, end_dt in intervals:
            gross = (end_dt - start_dt).total_seconds() / 3600.0
            allocated_break_hours = break_hours_row * (gross / total_gross_row_hours) if total_gross_row_hours else 0.0

            gross_nh = night_hours(start_dt, end_dt)
            gross_sh, sunday_dates = sunday_overlap(start_dt, end_dt, hd)
            gross_snh = sunday_night_hours(start_dt, end_dt, hd)
            row_holiday_dates = [d for d in sunday_dates if d in hd and d.weekday() != 6]

            # Κρατάμε τα πεδία net_* για συμβατότητα με τον υπόλοιπο κώδικα,
            # αλλά πλέον είναι ίσα με τις πραγματικές ώρες ωραρίου, χωρίς αφαίρεση διαλείμματος.
            net = gross
            net_nh = gross_nh
            net_sh = gross_sh
            net_snh = gross_snh

            summary.work_dates.add(start_dt.date())
            summary.gross_hours += gross
            summary.break_hours += allocated_break_hours
            summary.total_hours += gross
            summary.night_hours += gross_nh
            summary.sunday_hours += gross_sh
            summary.sunday_night_hours += gross_snh
            summary.sunday_dates.update(sunday_dates)
            summary.holiday_dates.update(row_holiday_dates)
            summary.work_rows += 1

            details.append(DetailRow(
                source_row=r,
                afm=afm,
                full_name=full_name,
                work_date=dt,
                break_minutes=allocated_break_hours * 60,
                apasxolisi=apasholisi,
                start_dt=start_dt,
                end_dt=end_dt,
                gross_hours=gross,
                net_hours=net,
                gross_night_hours=gross_nh,
                net_night_hours=net_nh,
                gross_sunday_hours=gross_sh,
                net_sunday_hours=net_sh,
                gross_sunday_night_hours=gross_snh,
                net_sunday_night_hours=net_snh,
                sunday_dates=sunday_dates,
                holiday_dates=row_holiday_dates,
            ))

    sorted_summaries = sorted(summaries.values(), key=lambda s: (s.last_name, s.first_name, s.afm))
    return sorted_summaries, details, excluded_rows, general_warnings


def autosize_columns(ws, max_width: int = 42):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = 10
        for row in range(1, min(ws.max_row, 200) + 1):
            value = ws.cell(row=row, column=col).value
            if value is not None:
                width = max(width, min(max_width, len(str(value)) + 2))
        ws.column_dimensions[letter].width = width


def write_summary_xlsx(output_path: str | Path, summaries: List[EmployeeSummary], details: List[DetailRow], excluded_rows: List[ExcludedRow], warnings: List[str]):
    output_path = Path(output_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Σύνοψη"

    headers = [
        "ΑΦΜ",
        "Ονοματεπώνυμο",
        "Ημέρες εργασίας",
        "Ώρες εργασίας",
        "Νυχτερινές ώρες",
        "Ημέρες εργασίας Κυριακής",
        "Ώρες Κυριακής",
        "Νυχτερινές ώρες Κυριακής",
        "Ώρες ωραρίου",
        "Διάλειμμα καταχωρημένο (ώρες - δεν αφαιρείται)",
        "Γραμμές/βάρδιες εργασίας",
        "Ημέρες άδειας που εξαιρέθηκαν",
        "Γραμμές άδειας που εξαιρέθηκαν",
        "Γραμμές μη εργασίας/ρεπό",
        "Παρατηρήσεις",
    ]
    ws.append(headers)
    for s in summaries:
        ws.append([
            s.afm,
            s.full_name,
            len(s.work_dates),
            round_hours(s.total_hours),
            round_hours(s.night_hours),
            len(s.sunday_dates),
            round_hours(s.sunday_hours),
            round_hours(s.sunday_night_hours),
            round_hours(s.gross_hours),
            round_hours(s.break_hours),
            s.work_rows,
            len(s.excluded_leave_dates),
            s.excluded_leave_rows,
            s.excluded_non_work_rows,
            leave_note(s.excluded_leave_rows)
            + holiday_note(s.holiday_dates)
            + ("; ".join(s.warnings[:3]) + (" ..." if len(s.warnings) > 3 else "")),
        ])

    detail_ws = wb.create_sheet("Αναλυτικές_Γραμμές")
    detail_ws.append([
        "Γραμμή αρχείου",
        "ΑΦΜ",
        "Ονοματεπώνυμο",
        "Ημερομηνία γραμμής",
        "Απασχόληση",
        "Έναρξη",
        "Λήξη",
        "Ώρες ωραρίου",
        "Διάλειμμα καταχωρημένο (λεπτά - δεν αφαιρείται)",
        "Ώρες εργασίας",
        "Μικτές νυχτερινές ώρες",
        "Νυχτερινές ώρες",
        "Ώρες ωραρίου Κυριακής",
        "Ώρες Κυριακής",
        "Μικτές νυχτερινές ώρες Κυριακής",
        "Νυχτερινές ώρες Κυριακής",
        "Κυριακές που αφορούνται",
        "Αργίες που αφορούνται",
    ])
    for d in details:
        detail_ws.append([
            d.source_row,
            d.afm,
            d.full_name,
            d.work_date.strftime("%d/%m/%Y"),
            d.apasxolisi,
            d.start_dt.strftime("%d/%m/%Y %H:%M"),
            d.end_dt.strftime("%d/%m/%Y %H:%M"),
            round_hours(d.gross_hours),
            round_hours(d.break_minutes),
            round_hours(d.net_hours),
            round_hours(d.gross_night_hours),
            round_hours(d.net_night_hours),
            round_hours(d.gross_sunday_hours),
            round_hours(d.net_sunday_hours),
            round_hours(d.gross_sunday_night_hours),
            round_hours(d.net_sunday_night_hours),
            ", ".join(x.strftime("%d/%m/%Y") for x in d.sunday_dates if x.weekday() == 6),
            ", ".join(x.strftime("%d/%m/%Y") for x in d.holiday_dates),
        ])

    excluded_ws = wb.create_sheet("Εξαιρούμενες_Γραμμές")
    excluded_ws.append(["Γραμμή αρχείου", "ΑΦΜ", "Ονοματεπώνυμο", "Ημερομηνία", "Απασχόληση", "Αιτία εξαίρεσης"])
    for e in excluded_rows:
        excluded_ws.append([
            e.source_row,
            e.afm,
            e.full_name,
            e.work_date.strftime("%d/%m/%Y") if e.work_date else "",
            e.apasxolisi,
            e.reason,
        ])

    checks_ws = wb.create_sheet("Έλεγχοι")
    total_leave_rows = sum(s.excluded_leave_rows for s in summaries)
    total_non_work_rows = sum(s.excluded_non_work_rows for s in summaries)
    total_work_rows = sum(s.work_rows for s in summaries)
    total_gross = sum(s.gross_hours for s in summaries)
    total_break = sum(s.break_hours for s in summaries)
    total_net = sum(s.total_hours for s in summaries)
    checks_rows = [
        ["Έλεγχος", "Τιμή"],
        ["Εργαζόμενοι", len(summaries)],
        ["Μετρημένες γραμμές/βάρδιες εργασίας", total_work_rows],
        ["Γραμμές άδειας που εξαιρέθηκαν", total_leave_rows],
        ["Γραμμές μη εργασίας/ρεπό που εξαιρέθηκαν", total_non_work_rows],
        ["Σύνολο ωρών εργασίας", round_hours(total_gross)],
        ["Σύνολο διαλείμματος καταχωρημένο ενημερωτικά", round_hours(total_break)],
        ["Σύνολο ωρών εργασίας χωρίς αφαίρεση διαλείμματος", round_hours(total_net)],
        ["Προειδοποιήσεις ανάγνωσης", len(warnings)],
    ]
    for row in checks_rows:
        checks_ws.append(row)

    info_ws = wb.create_sheet("Κανόνες")
    info_rows = [
        ["Κανόνας", "Τι εφαρμόζεται"],
        ["Νυχτερινές ώρες", "22:00 έως 06:00 της επόμενης ημέρας."],
        ["Κυριακή", "Ημερολογιακή Κυριακή 00:00 έως 24:00. Οι βάρδιες που περνούν τα μεσάνυχτα σπάνε χρονικά."],
        ["Αργίες", "Οι επιλεγμένες αργίες υπολογίζονται όπως η Κυριακή (00:00-24:00), στις ίδιες στήλες με την Κυριακή. Καταγράφονται στις Παρατηρήσεις."],
        ["Νυχτερινές ώρες Κυριακής", "Μόνο οι ώρες που είναι ταυτόχρονα Κυριακή/αργία και νυχτερινές: 00:00-06:00 και 22:00-24:00."],
        ["Ημέρες άδειας", "Γραμμές με ένδειξη άδειας, π.χ. Κανονική άδεια, δεν υπολογίζονται καθόλου σε ημέρες ή ώρες εργασίας, ακόμη κι αν περιέχουν ωράριο."],
        ["Μη εργασία/ρεπό", "Γραμμές όπως ΜΗ ΕΡΓΑΣΙΑ ή ΑΝΑΠΑΥΣΗ/ΡΕΠΟ δεν υπολογίζονται."],
        ["Διάλειμμα", "Το διάλειμμα δεν αφαιρείται από τις ώρες εργασίας. Θεωρείται ότι περιλαμβάνεται στο δηλωμένο ωράριο και εμφανίζεται μόνο ενημερωτικά."],
        ["Μη υπολογιζόμενα", "Δεν υπολογίζονται αποδοχές, προσαυξήσεις ή ασφαλιστικές εισφορές."],
        ["Μορφή αρχείου", "Απαιτούνται οι στήλες ΑΦΜ, Όνομα, Επώνυμο, Ημ/νια, Απασχόληση. Η στήλη Διάλειμμα αξιοποιείται όταν υπάρχει."],
    ]
    for row in info_rows:
        info_ws.append(row)
    if warnings:
        info_ws.append([])
        info_ws.append(["Προειδοποιήσεις ανάγνωσης", ""])
        for w in warnings[:200]:
            info_ws.append([w, ""])

    # Styling
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        autosize_columns(sheet)
        sheet.auto_filter.ref = sheet.dimensions

    # Number formats only on numeric hour columns.
    for col in [4, 5, 7, 8, 9, 10]:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col).number_format = "0.00"
    for col in [8, 9, 10, 11, 12, 13, 14, 15, 16]:
        for row in range(2, detail_ws.max_row + 1):
            detail_ws.cell(row=row, column=col).number_format = "0.00"
    for col in [2]:
        for row in range(2, checks_ws.max_row + 1):
            if isinstance(checks_ws.cell(row=row, column=col).value, float):
                checks_ws.cell(row=row, column=col).number_format = "0.00"

    wb.save(output_path)


def write_summary_csv(output_path: str | Path, summaries: List[EmployeeSummary]):
    output_path = Path(output_path)
    with output_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow([
            "ΑΦΜ",
            "Ονοματεπώνυμο",
            "Ημέρες εργασίας",
            "Ώρες εργασίας",
            "Νυχτερινές ώρες",
            "Ημέρες εργασίας Κυριακής",
            "Ώρες Κυριακής",
            "Νυχτερινές ώρες Κυριακής",
            "Ώρες ωραρίου",
            "Διάλειμμα καταχωρημένο (ώρες - δεν αφαιρείται)",
            "Γραμμές/βάρδιες εργασίας",
            "Ημέρες άδειας που εξαιρέθηκαν",
            "Γραμμές άδειας που εξαιρέθηκαν",
            "Γραμμές μη εργασίας/ρεπό",
            "Παρατηρήσεις",
        ])
        for s in summaries:
            writer.writerow([
                s.afm,
                s.full_name,
                len(s.work_dates),
                round_hours(s.total_hours),
                round_hours(s.night_hours),
                len(s.sunday_dates),
                round_hours(s.sunday_hours),
                round_hours(s.sunday_night_hours),
                round_hours(s.gross_hours),
                round_hours(s.break_hours),
                s.work_rows,
                len(s.excluded_leave_dates),
                s.excluded_leave_rows,
                s.excluded_non_work_rows,
                leave_note(s.excluded_leave_rows) + holiday_note(s.holiday_dates) + "; ".join(s.warnings),
            ])


# ------------------------- GUI -------------------------

GREEK_MONTHS = [
    "Ιανουάριος", "Φεβρουάριος", "Μάρτιος", "Απρίλιος", "Μάιος", "Ιούνιος",
    "Ιούλιος", "Αύγουστος", "Σεπτέμβριος", "Οκτώβριος", "Νοέμβριος", "Δεκέμβριος",
]


def launch_gui():
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk

    root = tk.Tk()
    root.title("Σύνοψη εργασίας / Κυριακών / νυχτερινών / αργιών από Excel")
    root.geometry("1320x780")

    today_date = date.today()
    state = {
        "summaries": [], "details": [], "excluded_rows": [], "warnings": [],
        "input_path": None,
        "holiday_dates": set(),
        "month": today_date.month,
        "year": today_date.year,
    }

    top = ttk.Frame(root, padding=10)
    top.pack(fill="x")

    title = ttk.Label(top, text="Σύνοψη εργαζομένων από αρχείο Excel", font=("Segoe UI", 14, "bold"))
    title.pack(anchor="w")
    subtitle = ttk.Label(
        top,
        text="Φόρτωσε το Excel με τις στήλες ΑΦΜ, Όνομα, Επώνυμο, Ημ/νια, Διάλειμμα, Απασχόληση. Οι άδειες/ρεπό εξαιρούνται. Το διάλειμμα δεν αφαιρείται, επειδή θεωρείται ότι περιλαμβάνεται στο δηλωμένο ωράριο εργασίας.",
        wraplength=1250,
    )
    subtitle.pack(anchor="w", pady=(3, 8))

    btn_frame = ttk.Frame(top)
    btn_frame.pack(fill="x")

    hol_frame = ttk.Frame(top)
    hol_frame.pack(fill="x", pady=(6, 0))

    ttk.Label(hol_frame, text="Μήνας υπολογισμού:").pack(side="left", padx=(0, 4))
    month_var = tk.StringVar(value=GREEK_MONTHS[today_date.month - 1])
    month_cb = ttk.Combobox(hol_frame, values=GREEK_MONTHS, state="readonly", width=14, textvariable=month_var)
    month_cb.pack(side="left", padx=(0, 8))
    ttk.Label(hol_frame, text="Έτος:").pack(side="left", padx=(0, 4))
    year_var = tk.IntVar(value=today_date.year)
    year_sb = ttk.Spinbox(hol_frame, from_=2000, to=2100, width=6, textvariable=year_var)
    year_sb.pack(side="left", padx=(0, 8))

    holiday_summary_var = tk.StringVar(value="Καμία αργία επιλεγμένη.")

    def open_holiday_dialog():
        try:
            year = int(year_var.get())
        except (tk.TclError, ValueError):
            messagebox.showerror("Σφάλμα", "Μη έγκυρο έτος.")
            return
        month = GREEK_MONTHS.index(month_var.get()) + 1
        state["month"], state["year"] = month, year

        dialog = tk.Toplevel(root)
        dialog.title(f"Αργίες — {GREEK_MONTHS[month - 1]} {year}")
        dialog.transient(root)
        dialog.grab_set()
        dialog.geometry("420x520")

        ttk.Label(dialog, text=f"Επίσημες αργίες {GREEK_MONTHS[month - 1]} {year}",
                  font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=12, pady=(10, 4))

        all_hols = greek_public_holidays(year)
        month_hols = sorted([(d, n) for d, n in all_hols.items() if d.month == month])

        official_vars: List[Tuple[date, tk.BooleanVar]] = []
        if not month_hols:
            ttk.Label(dialog, text="Δεν υπάρχουν επίσημες αργίες αυτόν τον μήνα.").pack(anchor="w", padx=12)
        else:
            for d, name in month_hols:
                checked = d in state["holiday_dates"] or not state["holiday_dates"]
                v = tk.BooleanVar(value=checked)
                ttk.Checkbutton(dialog, text=f"{d.strftime('%d/%m/%Y')} – {name}", variable=v).pack(anchor="w", padx=20)
                official_vars.append((d, v))

        ttk.Separator(dialog).pack(fill="x", pady=8, padx=10)
        ttk.Label(dialog, text="Επιπλέον αργίες (έως 3) — μορφή DD/MM",
                  font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=12, pady=(0, 4))

        extra_entries: List[tk.Entry] = []
        existing_extras = sorted([d for d in state["holiday_dates"]
                                  if d not in {x[0] for x in month_hols}])
        for i in range(3):
            row = ttk.Frame(dialog)
            row.pack(anchor="w", padx=20, pady=2)
            ttk.Label(row, text=f"Αργία {i + 1}:").pack(side="left", padx=(0, 6))
            e = ttk.Entry(row, width=10)
            if i < len(existing_extras):
                e.insert(0, existing_extras[i].strftime("%d/%m"))
            e.pack(side="left")
            extra_entries.append(e)

        msg_var = tk.StringVar(value="")
        ttk.Label(dialog, textvariable=msg_var, foreground="#c00").pack(anchor="w", padx=12, pady=(6, 0))

        def on_ok():
            selected: Set[date] = {d for d, v in official_vars if v.get()}
            for e in extra_entries:
                txt = e.get().strip()
                if not txt:
                    continue
                m = re.match(r"^\s*(\d{1,2})\s*[/.\-]\s*(\d{1,2})(?:\s*[/.\-]\s*(\d{2,4}))?\s*$", txt)
                if not m:
                    msg_var.set(f"Μη έγκυρη ημερομηνία: {txt} (DD/MM ή DD/MM/YYYY).")
                    return
                day = int(m.group(1))
                mon = int(m.group(2))
                yr = int(m.group(3)) if m.group(3) else year
                if yr < 100:
                    yr += 2000
                try:
                    selected.add(date(yr, mon, day))
                except ValueError:
                    msg_var.set(f"Μη έγκυρη ημερομηνία: {txt}.")
                    return
            state["holiday_dates"] = selected
            if selected:
                holiday_summary_var.set(
                    f"Αργίες ({len(selected)}): " +
                    ", ".join(d.strftime("%d/%m/%Y") for d in sorted(selected))
                )
            else:
                holiday_summary_var.set("Καμία αργία επιλεγμένη.")
            dialog.destroy()

        btns = ttk.Frame(dialog)
        btns.pack(side="bottom", fill="x", pady=10)
        ttk.Button(btns, text="OK", command=on_ok).pack(side="right", padx=10)
        ttk.Button(btns, text="Άκυρο", command=dialog.destroy).pack(side="right")

    ttk.Button(hol_frame, text="Αργίες μήνα...", command=open_holiday_dialog).pack(side="left", padx=(0, 8))
    ttk.Label(hol_frame, textvariable=holiday_summary_var, foreground="#1F4E78").pack(side="left")

    status_var = tk.StringVar(value="Δεν έχει φορτωθεί αρχείο.")

    columns = [
        "ΑΦΜ", "Ονοματεπώνυμο", "Ημέρες εργασίας", "Ώρες εργασίας", "Νυχτερινές",
        "Ημέρες Κυριακής", "Ώρες Κυριακής", "Νυχτ. Κυριακής", "Ώρες ωραρίου", "Διάλειμμα", "Βάρδιες", "Άδειες", "Ρεπό/μη εργ.", "Παρατηρήσεις"
    ]
    tree = ttk.Treeview(root, columns=columns, show="headings", height=24)
    widths = [90, 230, 105, 105, 115, 115, 130, 150, 95, 90, 70, 70, 90, 260]
    for col, width in zip(columns, widths):
        tree.heading(col, text=col)
        tree.column(col, width=width, anchor="center" if col not in {"Ονοματεπώνυμο", "Παρατηρήσεις"} else "w")

    vsb = ttk.Scrollbar(root, orient="vertical", command=tree.yview)
    hsb = ttk.Scrollbar(root, orient="horizontal", command=tree.xview)
    tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
    tree.pack(fill="both", expand=True, padx=10, pady=(0, 0))
    vsb.place(relx=1.0, rely=0.22, relheight=0.70, anchor="ne")
    hsb.pack(fill="x", padx=10)

    def refresh_table():
        for item in tree.get_children():
            tree.delete(item)
        for s in state["summaries"]:
            tree.insert("", "end", values=(
                s.afm,
                s.full_name,
                len(s.work_dates),
                f"{round_hours(s.total_hours):.2f}",
                f"{round_hours(s.night_hours):.2f}",
                len(s.sunday_dates),
                f"{round_hours(s.sunday_hours):.2f}",
                f"{round_hours(s.sunday_night_hours):.2f}",
                f"{round_hours(s.gross_hours):.2f}",
                f"{round_hours(s.break_hours):.2f}",
                s.work_rows,
                s.excluded_leave_rows,
                s.excluded_non_work_rows,
                leave_note(s.excluded_leave_rows)
                + holiday_note(s.holiday_dates)
                + ("; ".join(s.warnings[:2]) + (" ..." if len(s.warnings) > 2 else "")),
            ))
        status_var.set(
            f"Αρχείο: {state['input_path']} | Εργαζόμενοι: {len(state['summaries'])} | Μετρημένες βάρδιες: {len(state['details'])} | Εξαιρούμενες γραμμές: {len(state['excluded_rows'])} | Προειδοποιήσεις: {len(state['warnings'])}"
        )

    def load_file():
        path = filedialog.askopenfilename(
            title="Επιλογή αρχείου Excel",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")],
        )
        if not path:
            return
        try:
            summaries, details, excluded_rows, warnings = analyze_excel(path, holiday_dates=state["holiday_dates"])
            state["summaries"] = summaries
            state["details"] = details
            state["excluded_rows"] = excluded_rows
            state["warnings"] = warnings
            state["input_path"] = path
            refresh_table()
            messagebox.showinfo("Ολοκληρώθηκε", f"Φορτώθηκαν {len(summaries)} εργαζόμενοι, {len(details)} μετρημένες βάρδιες και {len(excluded_rows)} εξαιρούμενες γραμμές.")
        except Exception as exc:
            messagebox.showerror("Σφάλμα", str(exc))

    def export_xlsx():
        if not state["summaries"]:
            messagebox.showwarning("Δεν υπάρχει σύνοψη", "Φόρτωσε πρώτα ένα αρχείο Excel.")
            return
        default_name = "summary_ergasia_nyxta_kyriakes.xlsx"
        path = filedialog.asksaveasfilename(
            title="Αποθήκευση σύνοψης Excel",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel file", "*.xlsx")],
        )
        if not path:
            return
        try:
            write_summary_xlsx(path, state["summaries"], state["details"], state["excluded_rows"], state["warnings"])
            messagebox.showinfo("Έγινε", f"Αποθηκεύτηκε η σύνοψη:\n{path}")
        except Exception as exc:
            messagebox.showerror("Σφάλμα εξαγωγής", str(exc))

    def export_csv():
        if not state["summaries"]:
            messagebox.showwarning("Δεν υπάρχει σύνοψη", "Φόρτωσε πρώτα ένα αρχείο Excel.")
            return
        path = filedialog.asksaveasfilename(
            title="Αποθήκευση CSV",
            defaultextension=".csv",
            initialfile="summary_ergasia_nyxta_kyriakes.csv",
            filetypes=[("CSV file", "*.csv")],
        )
        if not path:
            return
        try:
            write_summary_csv(path, state["summaries"])
            messagebox.showinfo("Έγινε", f"Αποθηκεύτηκε το CSV:\n{path}")
        except Exception as exc:
            messagebox.showerror("Σφάλμα εξαγωγής", str(exc))

    load_btn = ttk.Button(btn_frame, text="1. Φόρτωση Excel", command=load_file)
    load_btn.pack(side="left", padx=(0, 8))
    export_btn = ttk.Button(btn_frame, text="2. Εξαγωγή σε Excel", command=export_xlsx)
    export_btn.pack(side="left", padx=(0, 8))
    csv_btn = ttk.Button(btn_frame, text="Εξαγωγή CSV", command=export_csv)
    csv_btn.pack(side="left", padx=(0, 8))

    status = ttk.Label(root, textvariable=status_var, padding=10)
    status.pack(fill="x")

    root.mainloop()


# ------------------------- Streamlit Web UI -------------------------

def is_running_under_streamlit() -> bool:
    """Detect Streamlit execution so the cloud app does not try to open Tkinter."""
    try:
        from streamlit.runtime.scriptrunner import get_script_run_ctx
        return get_script_run_ctx() is not None
    except Exception:
        return False


def launch_streamlit_app():
    import tempfile
    import streamlit as st

    st.set_page_config(page_title="Σύνοψη ωρών ΕΡΓΑΝΗ", page_icon="📊", layout="wide")
    st.title("Σύνοψη ωρών εργασίας από Excel")
    st.write(
        "Ανέβασε το αρχείο Excel στην ίδια μορφή και η εφαρμογή θα δημιουργήσει "
        "σύνοψη ανά εργαζόμενο. Το διάλειμμα δεν αφαιρείται από τις ώρες."
    )

    today_d = date.today()

    st.subheader("Μήνας υπολογισμού & αργίες")
    col_m, col_y = st.columns([2, 1])
    month = col_m.selectbox(
        "Μήνας υπολογισμού",
        list(range(1, 13)),
        format_func=lambda m: GREEK_MONTHS[m - 1],
        index=today_d.month - 1,
        key="calc_month",
    )
    year = col_y.number_input(
        "Έτος",
        min_value=2000, max_value=2100,
        value=today_d.year, step=1,
        key="calc_year",
    )
    year = int(year)

    all_hols = greek_public_holidays(year)
    month_hols = sorted([(d, n) for d, n in all_hols.items() if d.month == month])

    selected_official: Set[date] = set()
    with st.expander(
        f"Αργίες {GREEK_MONTHS[month - 1]} {year}",
        expanded=bool(month_hols),
    ):
        if not month_hols:
            st.caption("Δεν υπάρχουν επίσημες ελληνικές αργίες αυτόν τον μήνα.")
        else:
            st.caption(
                "Οι επιλεγμένες αργίες μετριούνται όπως η Κυριακή (00:00–24:00) "
                "και προστίθενται στις ώρες Κυριακής."
            )
            for d, name in month_hols:
                if st.checkbox(
                    f"{d.strftime('%d/%m/%Y')} – {name}",
                    value=True,
                    key=f"hol_off_{d.isoformat()}",
                ):
                    selected_official.add(d)

    extra: Set[date] = set()
    with st.expander("Προσθήκη επιπλέον αργίας (έως 3)", expanded=False):
        st.caption("Άφησέ τα κενά αν δεν χρειάζονται. Δέχεται οποιαδήποτε ημερομηνία.")
        ec1, ec2, ec3 = st.columns(3)
        for i, col in enumerate((ec1, ec2, ec3), start=1):
            v = col.date_input(
                f"Αργία {i}",
                value=None,
                key=f"extra_hol_{i}",
                format="DD/MM/YYYY",
            )
            if v:
                extra.add(v)

    holiday_dates: Set[date] = selected_official | extra
    if holiday_dates:
        st.info(
            "Επιλεγμένες αργίες: " +
            ", ".join(d.strftime("%d/%m/%Y") for d in sorted(holiday_dates))
        )

    uploaded = st.file_uploader("Αρχείο Excel (.xlsx)", type=["xlsx"])

    if uploaded is None:
        st.info("Ανέβασε ένα αρχείο .xlsx για να γίνει ο υπολογισμός.")
        return

    try:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir_path = Path(tmpdir)
            input_path = tmpdir_path / uploaded.name
            input_path.write_bytes(uploaded.getvalue())

            summaries, details, excluded_rows, warnings = analyze_excel(input_path, holiday_dates=holiday_dates)
            output_path = tmpdir_path / f"summary_{Path(uploaded.name).stem}.xlsx"
            write_summary_xlsx(output_path, summaries, details, excluded_rows, warnings)
            output_bytes = output_path.read_bytes()

        st.success("Ο υπολογισμός ολοκληρώθηκε.")

        total_work_rows = sum(s.work_rows for s in summaries)
        total_leave_rows = sum(s.excluded_leave_rows for s in summaries)
        total_non_work_rows = sum(s.excluded_non_work_rows for s in summaries)
        total_hours = sum(s.total_hours for s in summaries)
        total_night = sum(s.night_hours for s in summaries)
        total_sunday = sum(s.sunday_hours for s in summaries)
        total_sunday_night = sum(s.sunday_night_hours for s in summaries)

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Εργαζόμενοι", len(summaries))
        c2.metric("Βάρδιες εργασίας", total_work_rows)
        c3.metric("Ώρες εργασίας", round_hours(total_hours))
        c4.metric("Προειδοποιήσεις", len(warnings))

        c5, c6, c7, c8 = st.columns(4)
        c5.metric("Νυχτερινές ώρες", round_hours(total_night))
        c6.metric("Ώρες Κυριακής", round_hours(total_sunday))
        c7.metric("Νυχτερινές Κυριακής", round_hours(total_sunday_night))
        c8.metric("Γραμμές άδειας που εξαιρέθηκαν", total_leave_rows)

        preview_rows = []
        for s in summaries:
            preview_rows.append({
                "ΑΦΜ": s.afm,
                "Ονοματεπώνυμο": s.full_name,
                "Ημέρες εργασίας": len(s.work_dates),
                "Ώρες εργασίας": round_hours(s.total_hours),
                "Νυχτερινές ώρες": round_hours(s.night_hours),
                "Ημέρες Κυριακής/αργίας": len(s.sunday_dates),
                "Ώρες Κυριακής/αργίας": round_hours(s.sunday_hours),
                "Νυχτερινές Κυριακής/αργίας": round_hours(s.sunday_night_hours),
                "Αργίες": ", ".join(d.strftime("%d/%m") for d in sorted(s.holiday_dates)),
                "Γραμμές άδειας που εξαιρέθηκαν": s.excluded_leave_rows,
                "Γραμμές μη εργασίας/ρεπό": s.excluded_non_work_rows,
            })

        st.subheader("Προεπισκόπηση σύνοψης")
        st.dataframe(preview_rows, use_container_width=True, hide_index=True)

        st.download_button(
            label="Κατέβασμα αποτελέσματος σε Excel",
            data=output_bytes,
            file_name=f"summary_{Path(uploaded.name).stem}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        with st.expander("Κανόνες υπολογισμού"):
            st.write("• Οι γραμμές με άδεια εξαιρούνται πλήρως, ακόμη κι αν περιέχουν ωράριο.")
            st.write("• Οι γραμμές ΜΗ ΕΡΓΑΣΙΑ / ΑΝΑΠΑΥΣΗ / ΡΕΠΟ δεν υπολογίζονται.")
            st.write("• Το διάλειμμα δεν αφαιρείται από τις ώρες εργασίας.")
            st.write("• Τα νυχτερινά υπολογίζονται στο διάστημα 22:00–06:00.")
            st.write("• Η Κυριακή υπολογίζεται ημερολογιακά, από 00:00 έως 24:00.")
            st.write("• Οι επιλεγμένες αργίες υπολογίζονται όπως η Κυριακή και προστίθενται στις ίδιες στήλες.")

        if warnings:
            with st.expander("Προειδοποιήσεις ανάγνωσης"):
                for w in warnings[:200]:
                    st.warning(w)

    except Exception as exc:
        st.error("Δεν ολοκληρώθηκε η επεξεργασία του αρχείου.")
        st.exception(exc)


def main(argv: List[str]) -> int:
    # CLI mode για γρήγορη χρήση: python ergani_summary_app.py input.xlsx output.xlsx
    if len(argv) >= 3:
        input_path = Path(argv[1])
        output_path = Path(argv[2])
        summaries, details, excluded_rows, warnings = analyze_excel(input_path)
        if output_path.suffix.lower() == ".csv":
            write_summary_csv(output_path, summaries)
        else:
            write_summary_xlsx(output_path, summaries, details, excluded_rows, warnings)
        print(f"Ολοκληρώθηκε: {output_path}")
        return 0
    launch_gui()
    return 0


if __name__ == "__main__":
    if is_running_under_streamlit():
        launch_streamlit_app()
    else:
        raise SystemExit(main(sys.argv))
